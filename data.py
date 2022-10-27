'''
To generate the database, run the following commands in the terminal:
    > sqlite3 data.db < data.sql
    > python data.py

'''
import warnings
warnings.filterwarnings('ignore')

import sqlite3
import openpyxl as op


book = op.load_workbook("Mortality_incidence_sociodemographic_and_clinical_data_in_COVID19_patients.xlsx")

sheet = book.active 

#data columns dict
dc = {"LOS": 2, "Age": 26, "Death": 3, "Severity": 5, "Black": 6, "White": 7, "Asian": 8, "Latino": 9, 
      "MI": 10, "PVD": 11, "CHF": 12, "CVD": 13, "DEMENT": 14, "COPD": 15, "DM Complicated": 16, 
      "DM Simple": 17, "Renal Disease": 18, "All CNS": 19, "Pure CNS": 20, "Stroke": 21, "Seizure": 22}

#blood columns dict
bc = {"OsSats": 29, "Temp": 32, "MAP": 35, "Ddimer": 38, "Pits": 41, "INR": 44, "BUN": 47, "Creatinine": 50, "Sodium": 53, "Glucose": 56, 
      "AST": 59, "ALT": 62, "WBC": 65, "Lympho": 68, "IL6": 71, "Ferritin": 74, "CrctProtein": 77, "Procalcitonin": 80, "Troponin": 83}

# Extract Patient data and historical info

data = []
i = 0
id = 0

for row in sheet.iter_rows(min_row=2, max_row=4713):
    if i != 0:
        data.append(tuple(patient))
    i += 1
    id += 1
    patient = []
    patient.append(id)
    for column in dc.values():
        patient.append(row[column].value)

        
# Extract Patient Blood panel results

blood = []
k = 0
id = 0
patient_id = 0

for row in sheet.iter_rows(min_row=2, max_row=4713):
    if k != 0:
        blood.append(tuple(patient_blood))
    k += 1
    id += 1
    patient_id += 1
    patient_blood = []
    patient_blood.append(id)
    patient_blood.append(patient_id)
    for column in bc.values():
        patient_blood.append(row[column].value)


con = sqlite3.connect('data.db')

con.executemany('''
                INSERT INTO patient(id,length_of_stay, age, death, severity, black, white, asian, latino, 
                                    myocardial_infarction, peripheral_vascular_disease, congestive_heart_disease, 
                                    cardiovascular_disease, dementia,
                                    chronic_obstructive_pulmonary_disease,
                                    diabetes_mellitus_complicated,
                                    diabetes_mellitus_simple,
                                    renal_disease,
                                    all_central_nervous_system_disease,
                                    pure_central_nervous_system_disease,
                                    stroke,
                                    seizure) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ''', data);


con.executemany('''
                INSERT INTO blood(id, patient_id, oxygen_saturation, 
                                  temperature, mean_arterial_pressure, D_dimer, platelets, 
                                  international_normalized_ratio, blood_urea_nitrogen, creatinine, 
                                  sodium, glucose, aspartate_aminotransferase, alanine_aminotransferase, 
                                  white_blood_cell, lymphocytes, interleukin6, ferritin, C_reactive_protein, 
                                   procalcitonin ,troponin) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ''', blood);


con.commit();



